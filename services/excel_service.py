from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

REQUIRED_COLUMNS = ["Артикул", "Наименование", "EAN13", "Количество"]


def build_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Товары"
    ws.append(REQUIRED_COLUMNS)
    ws.append(["АРТ-001", "Товар 1", "", 1])
    ws.append(["АРТ-002", "Товар 2", "2000012345678", 2])
    fill = PatternFill("solid", fgColor="D9EAF7")
    font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def read_uploaded_excel(file) -> pd.DataFrame:
    df = pd.read_excel(file, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"В Excel отсутствуют колонки: {', '.join(missing)}")
    df = df[REQUIRED_COLUMNS].copy()
    df["Артикул"] = df["Артикул"].fillna("").astype(str).str.strip()
    df["Наименование"] = df["Наименование"].fillna("").astype(str).str.strip()
    df["EAN13"] = df["EAN13"].fillna("").astype(str).str.strip()
    df["Количество"] = df["Количество"].fillna("1").astype(str).str.strip()
    return df


def result_to_xlsx(df: pd.DataFrame) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Результат")
    return output.getvalue()
